"""
Metadata:
    File Name: upload_tasks_to_github.py
    Owner: Andrew (SilicaStormSiam)
    Purpose: Automates the creation of Excel file templates for GitHub Projects under the SilicaStormSiam user,
             parses tasks from these files, and performs a one-way sync to upload new tasks to the Initiating column
             of the corresponding GitHub Project, preventing duplicates using exact and fuzzy matching, with detailed logging.
    Version: 1.0.23
    Version Control:
        Version 1.0.23 (2025-07-25): Fixed GIPPING_TOKEN typo, set Excel column widths to 15.
        Version 1.0.22 (2025-07-25): Corrected GraphQL query to fetch user-level projects, added dir and type verification.
        Version 1.0.21 (2025-07-25): Ensured Notepad++ command precedes script, refined user-level GraphQL query.
        Version 1.0.20 (2025-07-24): Switched to user-level GraphQL query (partially applied).
    Change Log:
        2025-07-25 (v1.0.23): Corrected GIPPING_TOKEN to GITHUB_TOKEN, added column width adjustment to 15.
        2025-07-25 (v1.0.22): Fixed GraphQL query to target user-level projects, added dir and type for verification.
        2025-07-24 (v1.0.21): Placed Notepad++ command before script, refined GraphQL query for user-level projects.
        2025-07-24 (v1.0.20): Implemented GraphQL query for user-level projects (partially applied).
"""
import pandas as pd
from github import Github
from github.GithubException import GithubException
from gql import gql, Client
from gql.transport.requests import RequestsHTTPTransport
import os
from dotenv import load_dotenv
from datetime import datetime
from fuzzywuzzy import fuzz
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Configuration
ENV_FILE = r"C:\Users\Andrew\silicastormsiam\scripts\.env"
LOG_FILE = r"C:\Users\Andrew\silicastormsiam\scripts\github_upload.log"
REPO_NAME = "SilicaStormSiam/SilicaStormSiam"
DEFAULT_COLUMN = "Initiating"
SIMILARITY_THRESHOLD = 90
PROJECT_NUMBERS = [2, 3, 4, 5]  # Target project numbers
COLUMN_WIDTH = 15  # Set column width to 15

def log_message(message, level="INFO"):
    """Log message to file and console with timestamp and log level."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_entry = f"[{timestamp}] [{level}] {message}\n"
    print(log_entry.strip())
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(log_entry)

def get_github_projects(github_token):
    """Fetch GitHub Projects for the user using GraphQL API."""
    try:
        log_message("Fetching GitHub Projects for user 'SilicaStormSiam' using GraphQL.")
        transport = RequestsHTTPTransport(
            url="https://api.github.com/graphql",
            headers={"Authorization": f"Bearer {github_token}"},
            verify=True
        )
        client = Client(transport=transport, fetch_schema_from_transport=False)
        
        query = gql("""
        query($owner: String!) {
            user(login: $owner) {
                projectsV2(first: 100) {
                    nodes {
                        number
                        title
                        id
                    }
                }
            }
        }
        """)
        
        variables = {"owner": "SilicaStormSiam"}
        result = client.execute(query, variable_values=variables)
        
        projects = {
            node["number"]: {
                "name": node["title"],
                "id": node["id"]
            } for node in result["user"]["projectsV2"]["nodes"] if node["number"] in PROJECT_NUMBERS
        }
        
        if not projects:
            log_message(f"No projects found for user 'SilicaStormSiam' or project numbers {PROJECT_NUMBERS} not present.", "WARNING")
            return {}
        
        log_message(f"Retrieved {len(projects)} projects: {', '.join([f'ID {pid}: {p['name']}' for pid, p in projects.items()])}")
        return projects
    except Exception as e:
        log_message(f"Failed to retrieve projects: {e}", "ERROR")
        raise

def get_project_cards(github_token, project_id):
    """Fetch all cards in the project to determine column locations using GraphQL."""
    try:
        log_message(f"Fetching cards for project ID {project_id}.")
        transport = RequestsHTTPTransport(
            url="https://api.github.com/graphql",
            headers={"Authorization": f"Bearer {github_token}"},
            verify=True
        )
        client = Client(transport=transport, fetch_schema_from_transport=False)
        
        query = gql("""
        query($projectId: ID!) {
            node(id: $projectId) {
                ... on ProjectV2 {
                    items(first: 100) {
                        nodes {
                            id
                            content {
                                ... on Issue {
                                    id
                                    title
                                }
                            }
                            fieldValues(first: 10) {
                                nodes {
                                    ... on ProjectV2ItemFieldSingleSelectValue {
                                        name
                                        field {
                                            ... on ProjectV2SingleSelectField {
                                                id
                                                name
                                            }
                                        }
                                    }
                                }
                            }
                        }
                    }
                }
            }
        }
        """)
        
        variables = {"projectId": project_id}
        result = client.execute(query, variable_values=variables)
        
        card_locations = {}
        for item in result["node"]["items"]["nodes"]:
            if item["content"] and "id" in item["content"]:
                issue_id = item["content"]["id"]
                for field_value in item["fieldValues"]["nodes"]:
                    if field_value.get("field", {}).get("name") == "Status":
                        card_locations[issue_id] = field_value["name"]
        log_message(f"Retrieved {len(card_locations)} cards for project ID {project_id}.")
        return card_locations
    except Exception as e:
        log_message(f"Failed to retrieve project cards for ID {project_id}: {e}", "ERROR")
        return {}

def create_excel_file(project_id, project_name):
    """Create an Excel file template for a project if it doesn't exist, with column widths set to 15."""
    file_name = rf"C:\Users\Andrew\silicastormsiam\scripts\tasks_project{project_id}.xlsx"
    if not os.path.exists(file_name):
        columns = ["Task Name", "Description", "Process Group", "Assignee", "Due Date", "Labels"]
        df = pd.DataFrame(columns=columns)
        df.to_excel(file_name, index=False)
        # Set column widths to 15
        wb = Workbook()
        ws = wb.active
        for col_num, col_name in enumerate(columns, start=1):
            ws.column_dimensions[get_column_letter(col_num)].width = COLUMN_WIDTH
        wb.save(file_name)
        log_message(f"Created Excel file template '{file_name}' for Project ID {project_id} ({project_name}) with column widths set to {COLUMN_WIDTH}.")
    else:
        log_message(f"Excel file template '{file_name}' already exists for Project ID {project_id}.")
    return file_name

def get_initiating_column(github_token, project_id):
    """Retrieve the 'Initiating' column (status) for the project using GraphQL."""
    try:
        log_message(f"Fetching columns for project ID {project_id}.")
        transport = RequestsHTTPTransport(
            url="https://api.github.com/graphql",
            headers={"Authorization": f"Bearer {github_token}"},
            verify=True
        )
        client = Client(transport=transport, fetch_schema_from_transport=False)
        
        query = gql("""
        query($projectId: ID!) {
            node(id: $projectId) {
                ... on ProjectV2 {
                    fields(first: 10) {
                        nodes {
                            ... on ProjectV2SingleSelectField {
                                id
                                name
                                options {
                                    id
                                    name
                                }
                            }
                        }
                    }
                }
            }
        }
        """)
        
        variables = {"projectId": project_id}
        result = client.execute(query, variable_values=variables)
        
        for field in result["node"]["fields"]["nodes"]:
            if field["name"].lower() == "status":
                for option in field["options"]:
                    if option["name"].lower() == DEFAULT_COLUMN.lower():
                        log_message(f"Found 'Initiating' column for project ID {project_id}.")
                        return {"field_id": field["id"], "option_id": option["id"]}
        raise Exception(f"Column 'Initiating' not found in project ID {project_id}.")
    except Exception as e:
        log_message(f"Failed to retrieve columns for project ID {project_id}: {e}", "ERROR")
        raise

def create_issue(repo, task_name, description, github_token, project_id, project_name, column_info, assignees=None, due_date=None, labels=None):
    """Create a new GitHub issue and add it to the Initiating column using GraphQL."""
    try:
        issue_labels = ["task"]
        if labels:
            issue_labels.extend([label.strip() for label in labels.split(",") if label.strip()])
        log_message(f"Preparing labels for issue '{task_name}': {', '.join(issue_labels)}")
        
        assignee_list = [assignee.strip() for assignee in assignees.split(",") if assignee.strip()] if assignees else []
        log_message(f"Preparing assignees for issue '{task_name}': {', '.join(assignee_list) if assignee_list else 'None'}")
        
        issue = repo.create_issue(
            title=task_name,
            body=description,
            assignees=assignee_list,
            milestone=create_milestone(repo, due_date) if due_date else None,
            labels=issue_labels
        )
        log_message(f"Created issue '{task_name}' (ID {issue.id}) for Project ID {project_id}.")
        
        # Add issue to project using GraphQL
        transport = RequestsHTTPTransport(
            url="https://api.github.com/graphql",
            headers={"Authorization": f"Bearer {github_token}"},
            verify=True
        )
        client = Client(transport=transport, fetch_schema_from_transport=False)
        
        add_item_mutation = gql("""
        mutation($projectId: ID!, $contentId: ID!) {
            addProjectV2ItemById(input: {projectId: $projectId, contentId: $contentId}) {
                item {
                    id
                }
            }
        }
        """)
        
        update_field_mutation = gql("""
        mutation($projectId: ID!, $itemId: ID!, $fieldId: ID!, $optionId: ID!) {
            updateProjectV2ItemFieldValue(input: {projectId: $projectId, itemId: $itemId, fieldId: $fieldId, value: {singleSelectOptionId: $optionId}}) {
                projectV2Item {
                    id
                }
            }
        }
        """)
        
        # Add issue to project
        variables = {"projectId": project_id, "contentId": issue.id}
        result = client.execute(add_item_mutation, variable_values=variables)
        item_id = result["addProjectV2ItemById"]["item"]["id"]
        
        # Set status to Initiating
        variables = {
            "projectId": project_id,
            "itemId": item_id,
            "fieldId": column_info["field_id"],
            "optionId": column_info["option_id"]
        }
        client.execute(update_field_mutation, variable_values=variables)
        
        log_message(f"Added issue '{task_name}' to 'Initiating' in {project_name}.")
        return True
    except Exception as e:
        log_message(f"Failed to create issue '{task_name}' for Project ID {project_id}: {e}", "ERROR")
        return False

def create_milestone(repo, due_date):
    """Create or get a milestone for the due date."""
    if not due_date:
        log_message("No due date provided for milestone.")
        return None
    try:
        due_date = datetime.strptime(str(due_date), "%Y-%m-%d").date()
        log_message(f"Checking for milestone with due date {due_date}.")
        for milestone in repo.get_milestones():
            if milestone.due_on and milestone.due_on.date() == due_date:
                log_message(f"Found existing milestone 'Due {due_date}'.")
                return milestone
        milestone = repo.create_milestone(title=f"Due {due_date}", due_on=due_date)
        log_message(f"Created new milestone 'Due {due_date}'.")
        return milestone
    except (ValueError, GithubException) as e:
        log_message(f"Failed to create milestone for '{due_date}': {e}", "ERROR")
        return None

def main():
    try:
        log_message("Starting task upload process.")
        if not os.path.exists(ENV_FILE):
            log_message(f".env file not found at '{ENV_FILE}'. Please create it with GITHUB_TOKEN.", "ERROR")
            raise FileNotFoundError(f".env file missing at {ENV_FILE}")
        
        if not load_dotenv(ENV_FILE, override=True, encoding="utf-8"):
            log_message(f"Failed to load .env file at '{ENV_FILE}'. Ensure it contains a valid GITHUB_TOKEN.", "ERROR")
            raise ValueError(f"Invalid .env file at {ENV_FILE}")
        
        GITHUB_TOKEN = os.getenv("GITHUB_TOKEN")
        if not GITHUB_TOKEN or not GITHUB_TOKEN.startswith("ghp_") or len(GITHUB_TOKEN) < 40:
            log_message("Invalid or missing GITHUB_TOKEN in .env file. Expected format: ghp_ followed by 36-40 characters.", "ERROR")
            raise ValueError("Invalid GITHUB_TOKEN in .env")
        
        g = Github(GITHUB_TOKEN)
        try:
            repo = g.get_repo(REPO_NAME)
            log_message(f"Connected to repository '{REPO_NAME}'.")
        except GithubException as e:
            log_message(f"Failed to access repository '{REPO_NAME}': {e}. Check repository name and PAT permissions.", "ERROR")
            raise
        
        projects = get_github_projects(GITHUB_TOKEN)
        existing_issues = {}
        for issue in repo.get_issues(state="open", labels=["task"]):
            existing_issues[issue.title] = issue
        log_message(f"Retrieved {len(existing_issues)} existing open issues with 'task' label.")
        
        update_summary = {pid: {"total": 0, "successful": 0, "potential_duplicates": []} for pid in projects}
        
        for project_id, project in projects.items():
            try:
                project_name = project["name"]
                excel_file = create_excel_file(project_id, project_name)
                
                if not os.path.exists(excel_file):
                    log_message(f"Excel file '{excel_file}' not found for Project ID {project_id}. Skipping.", "WARNING")
                    continue
                
                log_message(f"Reading Excel file '{excel_file}' for Project ID {project_id}.")
                df = pd.read_excel(excel_file)
                required_columns = ["Task Name", "Description"]
                if not all(col in df.columns for col in required_columns):
                    log_message(f"File '{excel_file}' missing required columns: {', '.join(required_columns)}. Skipping.", "ERROR")
                    continue
                
                column_info = get_initiating_column(GITHUB_TOKEN, project["id"])
                card_locations = get_project_cards(GITHUB_TOKEN, project["id"])
                update_summary[project_id]["total"] = len(df)
                log_message(f"Processing {len(df)} tasks from '{excel_file}'.")
                
                for idx, row in df.iterrows():
                    task_name = row["Task Name"]
                    log_message(f"Checking task '{task_name}' (row {idx + 2}) for Project ID {project_id}.")
                    if task_name in existing_issues:
                        issue = existing_issues[task_name]
                        issue_id = str(issue.id)
                        column_name = card_locations.get(issue_id, "Unknown column")
                        log_message(f"Issue '{task_name}' already exists in {project_name} (Project ID {project_id}), located in '{column_name}'. Skipping to preserve existing data.")
                        continue
                    
                    # Check for similar titles
                    for existing_title in existing_issues:
                        similarity = fuzz.ratio(task_name.lower(), existing_title.lower())
                        if similarity >= SIMILARITY_THRESHOLD:
                            issue = existing_issues[existing_title]
                            issue_id = str(issue